"""One-shot script: dump hardcoded seed_data.py into data/seed_data.xlsx.

Run once to bootstrap the Excel pipeline:
    python seed_to_excel.py

Re-running overwrites the file. After this, edit the spreadsheet to update
hospital content; the kiosk will re-read it on next backend start (delete
kiosk.db too if you want existing rows refreshed).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import seed_data

LANGS = ("en", "te", "hi", "ta")
OUT = Path(__file__).parent / "data" / "seed_data.xlsx"


def _join(values) -> str:
    return "|".join(values or [])


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- departments ---
    ws = wb.active
    ws.title = "departments"
    headers = (
        ["id", "map_id", "floor"]
        + [f"name_{l}" for l in LANGS]
        + [f"directions_{l}" for l in LANGS]
        + [f"aliases_{l}" for l in LANGS]
    )
    ws.append(headers)
    for d in seed_data.DEPARTMENTS:
        ws.append([
            d["id"], d["map_id"], d["floor"],
            *(d["name"].get(l, "") for l in LANGS),
            *(d["directions"].get(l, "") for l in LANGS),
            *(_join(d["aliases"].get(l, [])) for l in LANGS),
        ])

    # --- doctors ---
    ws = wb.create_sheet("doctors")
    headers = (
        ["id"]
        + [f"name_{l}" for l in LANGS]
        + [f"specialty_{l}" for l in LANGS]
        + ["specialty_keys", "room", "slots_today"]
    )
    ws.append(headers)
    for doc in seed_data.DOCTORS:
        ws.append([
            doc["id"],
            *(doc["name"].get(l, "") for l in LANGS),
            *(doc["specialty"].get(l, "") for l in LANGS),
            _join(doc["specialty_keys"]),
            doc["room"],
            _join(doc["slots_today"]),
        ])

    # --- faqs ---
    ws = wb.create_sheet("faqs")
    headers = (
        [f"q_{l}" for l in LANGS]
        + [f"a_{l}" for l in LANGS]
        + ["tags"]
    )
    ws.append(headers)
    for faq in seed_data.FAQS:
        ws.append([
            *(faq["q"].get(l, "") for l in LANGS),
            *(faq["a"].get(l, "") for l in LANGS),
            faq["tags"],
        ])

    # --- emergency_keywords ---
    ws = wb.create_sheet("emergency_keywords")
    ws.append(["keyword"])
    for kw in seed_data.EMERGENCY_KEYWORDS:
        ws.append([kw])

    # --- visiting_hours ---
    ws = wb.create_sheet("visiting_hours")
    ws.append(["lang", "text"])
    for lang, text in seed_data.VISITING_HOURS.items():
        ws.append([lang, text])

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
