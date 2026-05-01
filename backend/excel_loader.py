"""Excel-driven seed loader.

Reads `data/seed_data.xlsx` and exposes the same module-level constants as
`seed_data.py` (DEPARTMENTS, DOCTORS, FAQS, EMERGENCY_KEYWORDS, VISITING_HOURS)
so non-developers can edit the spreadsheet to update kiosk content.

If the Excel file is missing or unreadable, falls back to the hardcoded
constants in `seed_data.py`.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import seed_data as _fallback

log = logging.getLogger("hospital-kiosk.excel_loader")

XLSX_PATH = Path(__file__).parent / "data" / "seed_data.xlsx"

LANGS = ("en", "te", "hi", "ta")


def _split_pipe(s: Any) -> list[str]:
    if s is None:
        return []
    text = str(s).strip()
    if not text:
        return []
    return [p.strip() for p in text.split("|") if p.strip()]


def _row_to_dict(headers: list[str], row: tuple) -> dict[str, Any]:
    return {h: row[i] for i, h in enumerate(headers) if i < len(row)}


def _read_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append(_row_to_dict(headers, row))
    return out


def _build_lang_dict(row: dict, prefix: str) -> dict[str, str]:
    return {lang: str(row.get(f"{prefix}_{lang}", "") or "") for lang in LANGS}


def _build_aliases(row: dict) -> dict[str, list[str]]:
    return {lang: _split_pipe(row.get(f"aliases_{lang}")) for lang in LANGS}


def _load_from_excel() -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)

    departments: list[dict] = []
    if "departments" in wb.sheetnames:
        for row in _read_sheet(wb["departments"]):
            departments.append({
                "id": str(row.get("id", "") or "").strip(),
                "map_id": str(row.get("map_id", "") or "").strip(),
                "floor": int(row.get("floor", 0) or 0),
                "name": _build_lang_dict(row, "name"),
                "directions": _build_lang_dict(row, "directions"),
                "aliases": _build_aliases(row),
            })

    doctors: list[dict] = []
    if "doctors" in wb.sheetnames:
        for row in _read_sheet(wb["doctors"]):
            doctors.append({
                "id": str(row.get("id", "") or "").strip(),
                "name": _build_lang_dict(row, "name"),
                "specialty": _build_lang_dict(row, "specialty"),
                "specialty_keys": _split_pipe(row.get("specialty_keys")),
                "room": str(row.get("room", "") or "").strip(),
                "slots_today": _split_pipe(row.get("slots_today")),
            })

    faqs: list[dict] = []
    if "faqs" in wb.sheetnames:
        for row in _read_sheet(wb["faqs"]):
            faqs.append({
                "q": _build_lang_dict(row, "q"),
                "a": _build_lang_dict(row, "a"),
                "tags": str(row.get("tags", "") or ""),
            })

    emergency: list[str] = []
    if "emergency_keywords" in wb.sheetnames:
        for row in _read_sheet(wb["emergency_keywords"]):
            kw = str(row.get("keyword", "") or "").strip()
            if kw:
                emergency.append(kw)

    visiting: dict[str, str] = {}
    if "visiting_hours" in wb.sheetnames:
        for row in _read_sheet(wb["visiting_hours"]):
            lang = str(row.get("lang", "") or "").strip()
            text = str(row.get("text", "") or "").strip()
            if lang and text:
                visiting[lang] = text

    wb.close()

    if not departments or not doctors:
        raise ValueError("Excel sheet missing departments/doctors data")

    return {
        "DEPARTMENTS": departments,
        "DOCTORS": doctors,
        "FAQS": faqs or _fallback.FAQS,
        "EMERGENCY_KEYWORDS": emergency or _fallback.EMERGENCY_KEYWORDS,
        "VISITING_HOURS": visiting or _fallback.VISITING_HOURS,
    }


def _load() -> dict[str, Any]:
    if XLSX_PATH.exists():
        try:
            data = _load_from_excel()
            log.info("Loaded seed data from %s", XLSX_PATH)
            return data
        except Exception as e:
            log.warning("Excel load failed (%s) — using hardcoded seed_data.py", e)
    else:
        log.info("Excel file %s not found — using hardcoded seed_data.py", XLSX_PATH)
    return {
        "DEPARTMENTS": _fallback.DEPARTMENTS,
        "DOCTORS": _fallback.DOCTORS,
        "FAQS": _fallback.FAQS,
        "EMERGENCY_KEYWORDS": _fallback.EMERGENCY_KEYWORDS,
        "VISITING_HOURS": _fallback.VISITING_HOURS,
    }


_data = _load()

DEPARTMENTS = _data["DEPARTMENTS"]
DOCTORS = _data["DOCTORS"]
FAQS = _data["FAQS"]
EMERGENCY_KEYWORDS = _data["EMERGENCY_KEYWORDS"]
VISITING_HOURS = _data["VISITING_HOURS"]
